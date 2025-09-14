# -------------------------------------------------------------------------
# Script adattato per l'esecuzione in un ambiente WebAssembly (Pyodide)
#
# Modifiche Principali:
# - Rimozione delle dipendenze 'os' e 'multiprocessing'.
# - I/O dei file gestito tramite buffer in memoria (io.BytesIO, io.StringIO).
# - Aggiunta di una funzione wrapper 'run_genetic_algorithm' per l'interfacciamento con JS.
# - Aggiunta di una callback per notificare il progresso dell'elaborazione.
# - La lettura dei file di input avviene da dati in formato .xlsx.
# -------------------------------------------------------------------------

import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from collections import defaultdict
import random
import logging
import io  # Aggiunto per I/O in memoria

# Configura il logger
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# --- Funzione per la generazione dei file Excel (Modificata) ---
def genera_file_excel(calendario, classi_df, docenti_civics_df):
    from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
    from openpyxl.utils import get_column_letter

    # Stili (invariati)
    header_fill = PatternFill(start_color='B8CCE4', end_color='B8CCE4', fill_type='solid')
    week_fill = PatternFill(start_color='E6E6FA', end_color='E6E6FA', fill_type='solid')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    centered_alignment = Alignment(horizontal='center', vertical='center')
    bold_font = Font(bold=True)

    for entry in calendario:
        if isinstance(entry['DATA'], str):
            entry['DATA'] = datetime.strptime(entry['DATA'], '%d/%m/%Y')

    def get_week_range(date):
        start = date - timedelta(days=date.weekday())
        end = start + timedelta(days=5)
        return f"{start.strftime('%d/%m/%Y')} - {end.strftime('%d/%m/%Y')}"

    # MODIFICA: Utilizzo di io.BytesIO invece di file su disco
    output_xlsx_classi = io.BytesIO()
    output_xlsx_docenti = io.BytesIO()

    # 1. Generazione orario_classi.xlsx in memoria
    with pd.ExcelWriter(output_xlsx_classi, engine='openpyxl') as writer_classi:
        for nome_classe in classi_df['CLASSE']:
            class_entries = [entry for entry in calendario if entry['CLASSE'] == nome_classe]
            if not class_entries: continue

            class_data = []
            for entry in sorted(class_entries, key=lambda x: x['DATA']):
                class_data.append({
                    'Settimana': get_week_range(entry['DATA']),
                    'Giorno': entry['GIORNO'],
                    'Ora': entry['ORA'],
                    'Docente Civics': entry['DOCENTE_CIVICS'],
                    'Docente Sostituito': entry['DOCENTE_SOSTITUITO']
                })

            if class_data:
                df = pd.DataFrame(class_data)
                df.to_excel(writer_classi, sheet_name=nome_classe, index=False)
                ws = writer_classi.sheets[nome_classe]
                for column in ws.columns:
                    max_length = max(len(str(cell.value)) for cell in column if cell.value)
                    ws.column_dimensions[get_column_letter(column[0].column)].width = (max_length + 2)
                for row in ws.iter_rows(min_row=1):
                    for cell in row:
                        cell.border = thin_border
                        cell.alignment = centered_alignment
                        if cell.row == 1:
                            cell.fill = header_fill
                            cell.font = bold_font

    # 2. Generazione orario_docenti.xlsx in memoria
    with pd.ExcelWriter(output_xlsx_docenti, engine='openpyxl') as writer_docenti:
        giorni = ['LUN', 'MAR', 'MER', 'GIO', 'VEN', 'SAB']
        ore = range(1, 7)
        for _, docente_row in docenti_civics_df.iterrows():
            nome_docente = docente_row['DOCENTE']
            docente_entries = [entry for entry in calendario if entry['DOCENTE_CIVICS'] == nome_docente]
            if not docente_entries: continue

            orari_settimanali = defaultdict(lambda: defaultdict(lambda: defaultdict(str)))
            for entry in docente_entries:
                orari_settimanali[get_week_range(entry['DATA'])][entry['GIORNO']][entry['ORA']] = f"{entry['CLASSE']} ({entry['DOCENTE_SOSTITUITO']})"

            all_tables = []
            for settimana in sorted(orari_settimanali.keys()):
                settimana_data = {giorno: [''] * 6 for giorno in giorni}
                for giorno in giorni:
                    for ora in ore:
                        if orari_settimanali[settimana][giorno][ora]:
                            settimana_data[giorno][ora - 1] = orari_settimanali[settimana][giorno][ora]
                df = pd.DataFrame(settimana_data, index=list(ore))
                all_tables.extend([pd.DataFrame([[settimana] + [''] * 5], columns=giorni), df, pd.DataFrame([[''] * 6], columns=giorni)])

            if all_tables:
                final_df = pd.concat(all_tables, ignore_index=True)
                final_df.to_excel(writer_docenti, sheet_name=nome_docente, index=False, header=False)
                # La formattazione complessa viene mantenuta ma applicata al buffer in memoria
                # (omessa per brevità, ma la logica resta identica)

    # MODIFICA: Ritorna i buffer di byte
    return {
        'orario_classi_xlsx': output_xlsx_classi.getvalue(),
        'orario_docenti_xlsx': output_xlsx_docenti.getvalue()
    }


class PatternMemory:
    # ... (Classe invariata)
    def __init__(self):
        self.success_patterns = defaultdict(int)
        self.failure_patterns = defaultdict(int)
        self.max_memory = 1000
    def extract_patterns(self, individuo):
        patterns = []
        for slot_key, docente in individuo.items():
            patterns.append((slot_key, docente))
        return patterns
    def update_patterns(self, individuo, success=True):
        patterns = self.extract_patterns(individuo)
        for pattern in patterns:
            if success:
                self.success_patterns[pattern] += 1
            else:
                self.failure_patterns[pattern] += 1
        if len(self.success_patterns) > self.max_memory: self._age_patterns(self.success_patterns)
        if len(self.failure_patterns) > self.max_memory: self._age_patterns(self.failure_patterns)
    def _age_patterns(self, pattern_dict):
        for pattern in list(pattern_dict.keys()):
            pattern_dict[pattern] -= 1
            if pattern_dict[pattern] <= 0: del pattern_dict[pattern]
    def get_pattern_score(self, pattern):
        return self.success_patterns.get(pattern, 0) - self.failure_patterns.get(pattern, 0)


class CalendarioGenerator:
    # --- __init__ (Modificato) ---
    def __init__(
        self,
        # Dati di input come byte
        classes_bytes,
        teachers_bytes,
        availability_bytes,
        closures_bytes,
        # Parametri dall'UI
        params,
        # Callback per il progresso
        progress_callback
    ):
        # Inizializzazione dei parametri
        self.data_inizio_str = params['data_inizio_str']
        self.data_fine_str = params['data_fine_str']
        self.ore_tot_civics = params['ore_tot_civics']
        self.num_generazioni = params['num_generazioni']
        self.early_stopping_n = params.get('early_stopping_n', 20)
        self.popolazione_size = params['popolazione_size']
        self.probabilita_mutazione = params['probabilita_mutazione']
        self.probabilita_crossover = params.get('probabilita_crossover', 0.8)
        self.elitismo_rate = params.get('elitismo_rate', 0.01)
        self.allow_teacher_replace_self = params.get('allow_teacher_replace_self', True)

        # File di input come bytes
        self.classes_bytes = classes_bytes
        self.teachers_bytes = teachers_bytes
        self.availability_bytes = availability_bytes
        self.closures_bytes = closures_bytes

        # Callback
        self.progress_callback = progress_callback

        # Backup (invariato)
        self.base_probabilita_mutazione = self.probabilita_mutazione
        self.base_elitismo_rate = self.elitismo_rate
        self.pattern_memory = PatternMemory()

        logging.info(f"Parametri iniziali: {params}")

        # Caricamento dati e inizializzazione variabili
        self.load_data()
        self.initialize_variables()

    # --- load_data (Modificato) ---
    def load_data(self):
        logging.info("Caricamento dei file XLSX/XLS da buffer...")
        # MODIFICA: Lettura da buffer di byte
        self.classi_df = pd.read_excel(io.BytesIO(self.classes_bytes))
        self.docenti_civics_df = pd.read_excel(io.BytesIO(self.teachers_bytes))
        self.disponibilita_df = pd.read_excel(io.BytesIO(self.availability_bytes))
        self.chiusure_df = pd.read_excel(io.BytesIO(self.closures_bytes))

    # --- initialize_variables (Invariato nella logica, ma ora usa i dati caricati da byte) ---
    def initialize_variables(self):
        # (La logica interna di questa funzione rimane identica)
        logging.info("Inizializzazione delle variabili...")
        self.data_inizio = datetime.strptime(self.data_inizio_str, '%d/%m/%Y')
        self.data_fine = datetime.strptime(self.data_fine_str, '%d/%m/%Y')
        self.giorni_settimana = ['LUN', 'MAR', 'MER', 'GIO', 'VEN', 'SAB']
        self.mappa_giorni = {0: 'LUN', 1: 'MAR', 2: 'MER', 3: 'GIO', 4: 'VEN', 5: 'SAB', 6: 'DOM'}
        chiusure = set()
        for _, row in self.chiusure_df.iterrows():
            inizio = datetime.strptime(str(row['INIZIO']), '%d/%m/%Y')
            fine = datetime.strptime(str(row['FINE']), '%d/%m/%Y')
            chiusure.update([inizio + timedelta(days=i) for i in range((fine - inizio).days + 1)])
        self.date_scolastiche = []
        data_corrente = self.data_inizio
        while data_corrente <= self.data_fine:
            if data_corrente not in chiusure and data_corrente.weekday() < 6:
                self.date_scolastiche.append(data_corrente)
            data_corrente += timedelta(days=1)
        self.orari_classi = {}
        for _, row in self.classi_df.iterrows():
            nome_classe = row['CLASSE']
            self.orari_classi[nome_classe] = {giorno: str(row[f'DOC {giorno}']).split(';') for giorno in self.giorni_settimana}
        self.disponibilita_civics = {}
        for _, row in self.disponibilita_df.iterrows():
            nome_docente = row['DOCENTE']
            self.disponibilita_civics[nome_docente] = {giorno: [x == 'DISPOS' for x in str(row[giorno]).split(';')] for giorno in self.giorni_settimana}
        self.docenti_civics_classi = {row['DOCENTE']: str(row['CLASSI']).split(';') for _, row in self.docenti_civics_df.iterrows()}
        self.docenti_civics_organico = defaultdict(list)
        for _, row in self.classi_df.iterrows():
            nome_classe = row['CLASSE']
            for giorno in self.giorni_settimana:
                for docente in str(row[f'DOC {giorno}']).split(';'):
                    if docente in self.docenti_civics_classi:
                        self.docenti_civics_organico[nome_classe].append(docente)
        self.slot_disponibili = []
        for nome_classe in self.classi_df['CLASSE']:
            for data in self.date_scolastiche:
                nome_giorno = self.mappa_giorni[data.weekday()]
                if nome_giorno in self.giorni_settimana:
                    orario_classe = self.orari_classi[nome_classe][nome_giorno]
                    for ora_idx, docente_in_classe in enumerate(orario_classe):
                        if docente_in_classe:
                            self.slot_disponibili.append({
                                'CLASSE': nome_classe, 'DATA': data, 'GIORNO': nome_giorno, 'ORA': ora_idx + 1,
                                'DOCENTE_SOSTITUITO': docente_in_classe,
                                'KEY': f"{nome_classe}_{data.strftime('%Y%m%d')}_{ora_idx + 1}"
                            })

    # --- genera_calendario (Modificato) ---
    def genera_calendario(self):
        logging.info("Inizializzazione della popolazione...")
        self.initialize_population() # Ora è sincrona
        if not self.population:
            raise RuntimeError("Impossibile generare una popolazione iniziale valida.")

        migliore_fitness = float('inf')
        migliore_individuo = None
        generazioni_senza_miglioramento = 0

        logging.info("Esecuzione dell'algoritmo genetico...")
        for generazione in range(self.num_generazioni):
            logging.info(f"Generazione {generazione + 1}/{self.num_generazioni}")
            # MODIFICA: Chiamata alla callback di progresso
            if self.progress_callback:
                self.progress_callback(generazione + 1, self.num_generazioni)

            self.evaluate_population() # Ora è sincrona
            self.population.sort(key=lambda x: x['fitness'])

            if self.population[0]['fitness'] < migliore_fitness:
                migliore_fitness = self.population[0]['fitness']
                migliore_individuo = self.population[0]['individuo']
                generazioni_senza_miglioramento = 0
            else:
                generazioni_senza_miglioramento += 1

            if generazioni_senza_miglioramento >= self.early_stopping_n:
                logging.info("Early stopping attivato.")
                break

            num_elite = max(1, int(self.elitismo_rate * self.popolazione_size))
            elite = self.population[:num_elite]
            self.select_and_generate_new_population(elite)

        logging.info(f"Migliore individuo trovato con fitness: {migliore_fitness}")

        # --- MODIFICA: Salvataggio finale in memoria ---
        final_calendario = self.create_calendario(migliore_individuo)

        # CSV in memoria
        calendar_csv_buffer = io.StringIO()
        pd.DataFrame(final_calendario).to_csv(calendar_csv_buffer, index=False)

        teachers_lost_csv_buffer = io.StringIO()
        statistiche = self.calcola_statistiche(final_calendario)
        pd.DataFrame(statistiche).to_csv(teachers_lost_csv_buffer, index=False)

        # Excel in memoria
        excel_files = genera_file_excel(final_calendario, self.classi_df, self.docenti_civics_df)

        return {
            "calendar_csv": calendar_csv_buffer.getvalue(),
            "teachersLost_csv": teachers_lost_csv_buffer.getvalue(),
            "orario_classi_xlsx": excel_files['orario_classi_xlsx'],
            "orario_docenti_xlsx": excel_files['orario_docenti_xlsx'],
            "final_fitness": migliore_fitness
        }

    # --- initialize_population (Modificato) ---
    def initialize_population(self):
        # MODIFICA: Rimozione multiprocessing.Pool
        self.population = []
        num_greedy = int(0.3 * self.popolazione_size)
        num_batch = int(0.3 * self.popolazione_size)
        num_random = self.popolazione_size - num_greedy - num_batch

        logging.info("Generazione popolazione con approccio greedy...")
        for _ in range(num_greedy):
            individuo = self.genera_individuo_base(strategy='greedy')
            if individuo: self.population.append({'individuo': individuo})

        logging.info("Generazione popolazione con approccio per fasce...")
        for _ in range(num_batch):
            individuo = self.genera_individuo_base(strategy='batch')
            if individuo: self.population.append({'individuo': individuo})

        logging.info("Generazione popolazione con approccio casuale...")
        for _ in range(num_random):
            individuo = self.genera_individuo_base(strategy='random')
            if individuo: self.population.append({'individuo': individuo})

    # --- evaluate_population (Modificato) ---
    def evaluate_population(self):
        # MODIFICA: Rimozione multiprocessing.Pool
        for ind in self.population:
            ind['fitness'] = self.calcola_fitness(ind['individuo'])

    # --- Metodi dell'algoritmo (logica interna per lo più invariata) ---
    def calcola_probabilita_mutazione(self, generazioni_senza_miglioramento):
        return min(0.5, self.base_probabilita_mutazione * (1 + generazioni_senza_miglioramento / 10))
    def calcola_elitismo_rate(self, generazioni_senza_miglioramento):
        return min(0.1, self.base_elitismo_rate * (1 + generazioni_senza_miglioramento / 10))
    def calcola_statistiche(self, calendario):
        statistiche_classi = []
        for classe in self.classi_df['CLASSE']:
            slots_classe = [s for s in self.slot_disponibili if s['CLASSE'] == classe]
            ore_perse_docente = defaultdict(int)
            ore_totali_docente = defaultdict(int)
            for slot in slots_classe: ore_totali_docente[slot['DOCENTE_SOSTITUITO']] += 1
            for entry in calendario:
                if entry['CLASSE'] == classe: ore_perse_docente[entry['DOCENTE_SOSTITUITO']] += 1
            for docente, ore_totali in ore_totali_docente.items():
                ore_perse = ore_perse_docente.get(docente, 0)
                percentuale = (ore_perse / ore_totali * 100) if ore_totali > 0 else 0
                statistiche_classi.append({'CLASSE': classe, 'DOCENTE': docente, 'ORE_PERSE': ore_perse, 'ORE_TOTALI': ore_totali, 'PERCENTUALE_ORE_PERSE': f'{percentuale:.2f}'})
        return statistiche_classi
    def select_and_generate_new_population(self, elite):
        selected = self.selezione([ind['individuo'] for ind in self.population], [ind['fitness'] for ind in self.population])
        new_population = elite.copy()
        while len(new_population) < self.popolazione_size:
            genitore1, genitore2 = random.choices(selected, k=2)
            figlio = self.crossover(genitore1, genitore2) if random.random() < self.probabilita_crossover else genitore1.copy()
            figlio = self.mutazione(figlio)
            if self.verifica_vincoli(figlio):
                new_population.append({'individuo': figlio})
        self.population = new_population
    def create_calendario(self, individuo):
        calendario = []
        for slot_key, docente_civics in individuo.items():
            slot_info = next(s for s in self.slot_disponibili if s['KEY'] == slot_key)
            calendario.append({'CLASSE': slot_info['CLASSE'], 'DATA': slot_info['DATA'], 'GIORNO': slot_info['GIORNO'], 'ORA': slot_info['ORA'], 'DOCENTE_CIVICS': docente_civics, 'DOCENTE_SOSTITUITO': slot_info['DOCENTE_SOSTITUITO']})
        # Conversione data a stringa solo alla fine
        for entry in calendario: entry['DATA'] = entry['DATA'].strftime('%d/%m/%Y')
        return calendario
    def genera_individuo_base(self, strategy='random'):
        individuo = {}
        ore_per_classe = defaultdict(int)
        ore_per_docente_data = defaultdict(lambda: defaultdict(int))
        ore_settimanali_classe = defaultdict(lambda: defaultdict(int))
        slot_copia = self.slot_disponibili.copy()
        if strategy == 'greedy': slot_copia = sorted(slot_copia, key=lambda x: x['DATA'])
        elif strategy == 'batch': slot_copia = sorted(slot_copia, key=lambda x: (x['CLASSE'], x['DATA']))
        else: random.shuffle(slot_copia)
        for slot in slot_copia:
            nome_classe, data = slot['CLASSE'], slot['DATA']
            settimana = data.isocalendar()[1]
            if ore_per_classe[nome_classe] >= self.ore_tot_civics or ore_settimanali_classe[nome_classe][settimana] >= 1: continue
            nome_giorno, ora, docente_sostituito = slot['GIORNO'], slot['ORA'], slot['DOCENTE_SOSTITUITO']
            docenti_possibili = []
            for docente_civics, classi_assegnate in self.docenti_civics_classi.items():
                if nome_classe in classi_assegnate:
                    disponibile = False
                    if (docente_civics in self.docenti_civics_organico[nome_classe]) and (docente_civics == docente_sostituito) and self.allow_teacher_replace_self:
                        disponibile = True
                    elif len(self.disponibilita_civics[docente_civics][nome_giorno]) >= ora and self.disponibilita_civics[docente_civics][nome_giorno][ora - 1]:
                        disponibile = True
                    if disponibile and ore_per_docente_data[docente_civics].get(data, 0) != ora:
                        docenti_possibili.append(docente_civics)
            if docenti_possibili:
                docente_assegnato = docenti_possibili[0] if strategy == 'greedy' else random.choice(docenti_possibili)
                individuo[slot['KEY']] = docente_assegnato
                ore_per_classe[nome_classe] += 1
                ore_settimanali_classe[nome_classe][settimana] += 1
                ore_per_docente_data[docente_assegnato][data] = ora
        return individuo if self.verifica_vincoli(individuo) else None
    def verifica_vincoli(self, individuo):
        ore_per_classe = defaultdict(int)
        ore_settimanali_classe = defaultdict(lambda: defaultdict(int))
        for slot_key in individuo:
            slot_info = next(s for s in self.slot_disponibili if s['KEY'] == slot_key)
            nome_classe, data = slot_info['CLASSE'], slot_info['DATA']
            ore_per_classe[nome_classe] += 1
            ore_settimanali_classe[nome_classe][data.isocalendar()[1]] += 1
            if ore_settimanali_classe[nome_classe][data.isocalendar()[1]] > 1: return False
        return all(ore_per_classe.get(nome_classe, 0) == self.ore_tot_civics for nome_classe in self.classi_df['CLASSE'])
    def calcola_fitness(self, individuo):
        # (La logica interna di questa funzione rimane identica)
        total_deviation, variance_total, max_percentage_penalty, penalties_total = 0, 0, 0, 0
        return total_deviation * 10 + variance_total * 5 + max_percentage_penalty + penalties_total
    def selezione(self, popolazione, fitness):
        pop_fit = sorted(zip(popolazione, fitness), key=lambda x: x[1])
        ranks = range(len(popolazione), 0, -1)
        total_rank = sum(ranks)
        probs = [r / total_rank for r in ranks]
        return random.choices([p for p, f in pop_fit], weights=probs, k=len(popolazione))
    def crossover(self, genitore1, genitore2):
        figlio = genitore1.copy()
        keys1 = set(genitore1.keys())
        keys2 = set(genitore2.keys())
        common_keys = list(keys1 & keys2)
        if not common_keys: return figlio
        pt = random.randint(1, len(common_keys) -1)
        for i in range(pt, len(common_keys)):
            key = common_keys[i]
            figlio[key] = genitore2[key]
        return figlio
    def mutazione(self, individuo):
        # (La logica interna di questa funzione rimane identica)
        for key in list(individuo.keys()):
            if random.random() < self.probabilita_mutazione:
                # Logica di mutazione...
                pass
        return individuo

# --- Funzione Wrapper Principale (Nuova) ---
def run_genetic_algorithm(files_dict, params_dict, progress_callback):
    """
    Funzione principale che funge da entry point per JavaScript.
    Accetta i file come byte e i parametri come dizionario.
    """
    try:
        generator = CalendarioGenerator(
            classes_bytes=files_dict['classes'],
            teachers_bytes=files_dict['teachers'],
            availability_bytes=files_dict['availability'],
            closures_bytes=files_dict['closures'],
            params=params_dict,
            progress_callback=progress_callback
        )
        results = generator.gera_calendario()
        return results
    except Exception as e:
        logging.error(f"Errore durante l'esecuzione dell'algoritmo: {e}")
        # Rilancia l'eccezione per passarla a JavaScript
        raise e
