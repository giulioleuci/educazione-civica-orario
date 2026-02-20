# -------------------------------------------------------------------------
# Questo script utilizza un algoritmo genetico per generare un calendario di sostituzioni
# di docenti di educazione civica, in modo da distribuire in maniera più omogenea
# le ore perse tra i docenti e le classi.
#
# **Obiettivi:**
# L'obiettivo è quello di creare un calendario di sostituzioni che, partendo dai docenti
# di educazione civica disponibili, copra le ore di lezione previste affinché ogni classe
# raggiunga le ore totali di educazione civica richieste (ore_tot_civics), distribuendo
# in modo equilibrato le ore perse per docente e minimizzando le deviazioni di vario tipo.
#
# **File di Input:**
#  - classes.csv: contiene l'elenco delle classi e gli orari settimanali dei docenti.
#  - civics_teachers.csv: elenco dei docenti di educazione civica e le classi a loro assegnate.
#  - availability.csv: disponibilità oraria dei docenti di educazione civica (DISPOS o meno).
#  - closures.csv: elenca i periodi di chiusura della scuola.
#
# **File di Output (generati nella cartella di output impostata in cartella_output):**
#  - calendar.csv: calendario risultante con data, giorno, ora, docente civics e docente sostituito.
#  - teachersLost.csv: statistiche per classe e docente, con percentuale di ore perse.
#  - orario_classi.xlsx: un file Excel per ogni classe, con una vista settimanale delle sostituzioni.
#  - orario_docenti.xlsx: un file Excel per ogni docente di educazione civica, con una vista settimanale.
#
# **Organizzazione dell'Output:**
#  Oltre ai file finali, ad ogni generazione dell'algoritmo genetico verrà creata una cartella
#  denominata "generation_X" (dove X è il numero della generazione), contenente:
#   - calendar.csv: il calendario di quella generazione
#   - teachersLost.csv: le statistiche per quella generazione
#   - orario_classi.xlsx e orario_docenti.xlsx: i file Excel con la pianificazione per classi e docenti
#
# In questo modo potremo monitorare il progresso dell'algoritmo e verificare come le soluzioni
# si evolvono nel tempo.
#
# -------------------------------------------------------------------------


import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from collections import defaultdict
import os
import random
import multiprocessing
import logging

# Configura il logger per informazioni sull'esecuzione
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

def genera_file_excel(calendario, classi_df, docenti_civics_df, cartella_output):
    # Questa funzione genera due file Excel:
    # 1. orario_classi.xlsx: un foglio per ogni classe, con le sostituzioni settimanali
    # 2. orario_docenti.xlsx: un foglio per ogni docente, con le ore settimanali su righe
    # Commenti inline per guidare passo-passo

    from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
    from openpyxl.utils import get_column_letter

    # Stili personalizzati per le celle Excel
    header_fill = PatternFill(start_color='B8CCE4', end_color='B8CCE4', fill_type='solid')
    week_fill = PatternFill(start_color='E6E6FA', end_color='E6E6FA', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    centered_alignment = Alignment(horizontal='center', vertical='center')
    bold_font = Font(bold=True)

    # Converte le date in datetime, se stringhe
    for entry in calendario:
        if isinstance(entry['DATA'], str):
            entry['DATA'] = datetime.strptime(entry['DATA'], '%d/%m/%Y')

    # Funzione per calcolare il range settimanale (lun-sab)
    def get_week_range(date):
        start = date - timedelta(days=date.weekday())  # Lunedì
        end = start + timedelta(days=5)  # Sabato
        return f"{start.strftime('%d/%m/%Y')} - {end.strftime('%d/%m/%Y')}"

    # ---------------------------------------
    # 1. Generazione orario_classi.xlsx
    # ---------------------------------------
    writer_classi = pd.ExcelWriter(os.path.join(cartella_output, 'orario_classi.xlsx'), engine='openpyxl')

    # Per ogni classe, estraiamo le entries corrispondenti e creiamo un foglio
    for nome_classe in classi_df['CLASSE']:
        class_entries = [entry for entry in calendario if entry['CLASSE'] == nome_classe]

        class_data = []
        for entry in sorted(class_entries, key=lambda x: x['DATA']):
            class_data.append({
                'Settimana': get_week_range(entry['DATA']),
                'Giorno': entry['GIORNO'],
                'Ora': entry['ORA'],
                'Docente Civics': entry['DOCENTE_CIVICS'],
                'Docente Sostituito': entry['DOCENTE_SOSTITUITO']
            })

        if class_data:
            df = pd.DataFrame(class_data)
            df.to_excel(writer_classi, sheet_name=nome_classe, index=False)

            ws = writer_classi.sheets[nome_classe]

            # Adatta la larghezza delle colonne in base al contenuto
            for column in ws.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

            # Formattazione delle celle: bordi, allineamento e header
            for row in ws.iter_rows(min_row=1):
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = centered_alignment
                    if cell.row == 1:  # Header
                        cell.fill = header_fill
                        cell.font = bold_font

    writer_classi.close()

    # ---------------------------------------
    # 2. Generazione orario_docenti.xlsx
    # ---------------------------------------
    writer_docenti = pd.ExcelWriter(os.path.join(cartella_output, 'orario_docenti.xlsx'), engine='openpyxl')

    giorni = ['LUN', 'MAR', 'MER', 'GIO', 'VEN', 'SAB']
    ore = range(1, 7)  # Ore di lezione: 1-6

    # Per ogni docente di civics creiamo uno sheet con la suddivisione settimanale
    for _, docente_row in docenti_civics_df.iterrows():
        nome_docente = docente_row['DOCENTE']
        docente_entries = [entry for entry in calendario if entry['DOCENTE_CIVICS'] == nome_docente]

        # Struttura dati per salvare le lezioni per settimana, giorno, ora
        orari_settimanali = defaultdict(lambda: defaultdict(lambda: defaultdict(str)))

        # Riempimento delle strutture con le ore del docente
        for entry in docente_entries:
            settimana = get_week_range(entry['DATA'])
            giorno = entry['GIORNO']
            ora = entry['ORA']
            nome_classe = entry['CLASSE']
            docente_sostituito = entry['DOCENTE_SOSTITUITO']
            orari_settimanali[settimana][giorno][ora] = f"{nome_classe} ({docente_sostituito})"

        # Creiamo un DataFrame concatenando blocchi per ogni settimana
        all_tables = []
        for settimana in sorted(orari_settimanali.keys()):
            settimana_data = {giorno: [''] * 6 for giorno in giorni}

            for giorno in giorni:
                for ora in ore:
                    if orari_settimanali[settimana][giorno][ora]:
                        settimana_data[giorno][ora - 1] = orari_settimanali[settimana][giorno][ora]

            df = pd.DataFrame(settimana_data, index=list(ore))
            title_df = pd.DataFrame([[settimana] + [''] * (len(giorni) - 1)], columns=giorni)
            empty_df = pd.DataFrame([[''] * len(giorni)], columns=giorni)

            all_tables.extend([title_df, df, empty_df])

        # Scrittura su file Excel
        if all_tables:
            final_df = pd.concat(all_tables)
            final_df.to_excel(writer_docenti, sheet_name=nome_docente, index=False)

            ws = writer_docenti.sheets[nome_docente]

            # Imposta larghezza colonne
            for col in range(1, len(giorni) + 1):
                ws.column_dimensions[get_column_letter(col)].width = 20

            # Formattazione celle
            row_num = 1
            # Ciclo per formattare le righe relative alle settimane, alle ore e alle righe vuote
            while row_num <= ws.max_row:
                if isinstance(ws.cell(row=row_num, column=1).value, str) and \
                        '-' in str(ws.cell(row=row_num, column=1).value):
                    # Riga della settimana
                    for col in range(1, len(giorni) + 1):
                        cell = ws.cell(row=row_num, column=col)
                        cell.fill = week_fill
                        cell.border = thin_border
                        cell.font = bold_font
                        cell.alignment = centered_alignment
                    row_num += 1

                    # Riga con l'orario settimanale
                    for hour_row in range(6):
                        for col in range(1, len(giorni) + 1):
                            cell = ws.cell(row=row_num + hour_row, column=col)
                            cell.border = thin_border
                            cell.alignment = centered_alignment
                            if col == 1:
                                cell.fill = header_fill
                                cell.font = bold_font

                    # Salto la riga vuota successiva
                    row_num += 7
                else:
                    row_num += 1

    writer_docenti.close()


class PatternMemory:
    # Classe per gestire pattern di successo e fallimento, utile per memorizzare
    # combinazioni particolari di assegnazione nella popolazione
    def __init__(self):
        self.success_patterns = defaultdict(int)
        self.failure_patterns = defaultdict(int)
        self.max_memory = 1000  # Max pattern memorizzabili

    def extract_patterns(self, individuo):
        # Estrae i pattern dall'individuo
        patterns = []
        for slot_key, docente in individuo.items():
            patterns.append((slot_key, docente))
        return patterns

    def update_patterns(self, individuo, success=True):
        # Aggiorna i pattern di successo/fallimento in base all'esito
        patterns = self.extract_patterns(individuo)
        for pattern in patterns:
            if success:
                self.success_patterns[pattern] += 1
            else:
                self.failure_patterns[pattern] += 1

        # Rimozione graduale dei pattern troppo vecchi/meno rilevanti
        if len(self.success_patterns) > self.max_memory:
            self._age_patterns(self.success_patterns)
        if len(self.failure_patterns) > self.max_memory:
            self._age_patterns(self.failure_patterns)

    def _age_patterns(self, pattern_dict):
        # Diminuisce il "peso" di ogni pattern, eliminando quelli irrilevanti
        for pattern in list(pattern_dict.keys()):
            pattern_dict[pattern] -= 1
            if pattern_dict[pattern] <= 0:
                del pattern_dict[pattern]

    def get_pattern_score(self, pattern):
        # Calcolo di un punteggio pattern: successi - fallimenti
        success = self.success_patterns.get(pattern, 0)
        failure = self.failure_patterns.get(pattern, 0)
        return success - failure


class CalendarioGenerator:
    # Classe principale che gestisce l'esecuzione dell'algoritmo genetico
    def __init__(
        self,
        num_varianti=1,
        data_inizio_str='02/12/2024',
        data_fine_str='10/06/2025',
        ore_tot_civics=27,
        cartella_output="CALENDARIO_GENERATO",
        num_generazioni=200,
        early_stopping_n=20,
        popolazione_size=200,
        probabilita_mutazione=0.2,
        probabilita_crossover=0.8,
        elitismo_rate=0.01,
        num_cores=4,
        allow_teacher_replace_self=True
    ):
        # Inizializzazione dei parametri
        self.num_varianti = num_varianti
        self.data_inizio_str = data_inizio_str
        self.data_fine_str = data_fine_str
        self.ore_tot_civics = ore_tot_civics
        self.cartella_output = cartella_output
        self.num_generazioni = num_generazioni
        self.early_stopping_n = early_stopping_n
        self.popolazione_size = popolazione_size
        self.probabilita_mutazione = probabilita_mutazione
        self.probabilita_crossover = probabilita_crossover
        self.elitismo_rate = elitismo_rate
        self.num_cores = num_cores
        self.allow_teacher_replace_self = allow_teacher_replace_self

        # Backup degli hyperparams di base
        self.base_probabilita_mutazione = probabilita_mutazione
        self.base_probabilita_crossover = probabilita_crossover
        self.base_elitismo_rate = elitismo_rate
        self.base_early_stopping_n = early_stopping_n

        self.hyperparams = {
            'probabilita_mutazione': self.probabilita_mutazione,
            'probabilita_crossover': self.probabilita_crossover,
            'elitismo_rate': self.elitismo_rate
        }

        self.pattern_memory = PatternMemory()

        # Stampa parametri di inizializzazione
        print("Parametri iniziali:")
        print(f"num_varianti = {self.num_varianti}")
        print(f"data_inizio_str = {self.data_inizio_str}")
        print(f"data_fine_str = {self.data_fine_str}")
        print(f"ore_tot_civics = {self.ore_tot_civics}")
        print(f"cartella_output = {self.cartella_output}")
        print(f"num_generazioni = {self.num_generazioni}")
        print(f"early_stopping_n = {self.early_stopping_n}")
        print(f"popolazione_size = {self.popolazione_size}")
        print(f"probabilita_mutazione = {self.probabilita_mutazione}")
        print(f"probabilita_crossover = {self.probabilita_crossover}")
        print(f"elitismo_rate = {self.elitismo_rate}")
        print(f"num_cores = {self.num_cores}")
        print(f"allow_teacher_replace_self = {self.allow_teacher_replace_self}")

        # Caricamento dati e inizializzazione variabili
        self.load_data()
        self.initialize_variables()

    def calcola_probabilita_mutazione(self, generazioni_senza_miglioramento):
        # Aumenta gradualmente la probabilità di mutazione se non c'è miglioramento
        base_prob = self.base_probabilita_mutazione
        return min(0.5, base_prob * (1 + generazioni_senza_miglioramento / 10))

    def calcola_elitismo_rate(self, generazioni_senza_miglioramento):
        # Aumenta gradualmente il tasso di elitismo se non c'è miglioramento
        base_rate = self.base_elitismo_rate
        return min(0.1, base_rate * (1 + generazioni_senza_miglioramento / 10))

    def load_data(self):
        # Caricamento dati da file CSV
        logging.info("Caricamento dei file CSV...")
        self.classi_df = pd.read_csv('classes.csv')
        self.docenti_civics_df = pd.read_csv('civics_teachers.csv')
        self.disponibilita_df = pd.read_csv('availability.csv')
        self.chiusure_df = pd.read_csv('closures.csv')

    def initialize_variables(self):
        logging.info("Inizializzazione delle variabili...")
        self.data_inizio = datetime.strptime(self.data_inizio_str, '%d/%m/%Y')
        self.data_fine = datetime.strptime(self.data_fine_str, '%d/%m/%Y')
        self.giorni_settimana = ['LUN', 'MAR', 'MER', 'GIO', 'VEN', 'SAB']
        self.mappa_giorni = {0: 'LUN', 1: 'MAR', 2: 'MER', 3: 'GIO', 4: 'VEN', 5: 'SAB', 6: 'DOM'}

        # Creazione della lista di date scolastiche escludendo i giorni di chiusura
        logging.info("Creazione della lista delle date scolastiche escludendo le chiusure...")
        chiusure = set()
        for _, row in self.chiusure_df.iterrows():
            inizio = datetime.strptime(row['INIZIO'], '%d/%m/%Y')
            fine = datetime.strptime(row['FINE'], '%d/%m/%Y')
            chiusure.update([inizio + timedelta(days=i) for i in range((fine - inizio).days + 1)])

        self.date_scolastiche = []
        data_corrente = self.data_inizio
        while data_corrente <= self.data_fine:
            if data_corrente not in chiusure and data_corrente.weekday() < 6:
                self.date_scolastiche.append(data_corrente)
            data_corrente += timedelta(days=1)

        logging.info(f"Numero totale di giorni scolastici: {len(self.date_scolastiche)}")

        # Parsing orari delle classi
        logging.info("Parsing degli orari delle classi...")
        self.orari_classi = {}
        for _, row in self.classi_df.iterrows():
            nome_classe = row['CLASSE']
            self.orari_classi[nome_classe] = {}
            for giorno in self.giorni_settimana:
                colonna_doc = f'DOC {giorno}'
                lista_docenti = row[colonna_doc].split(';')
                self.orari_classi[nome_classe][giorno] = lista_docenti

        # Parsing disponibilità docenti civics
        logging.info("Parsing della disponibilità dei docenti di educazione civica...")
        self.disponibilita_civics = {}
        for _, row in self.disponibilita_df.iterrows():
            nome_docente = row['DOCENTE']
            self.disponibilita_civics[nome_docente] = {}
            for giorno in self.giorni_settimana:
                disponibilita_stringa = row[giorno]
                disponibilita_lista = disponibilita_stringa.split(';')
                disponibilita_bool = [x == 'DISPOS' for x in disponibilita_lista]
                self.disponibilita_civics[nome_docente][giorno] = disponibilita_bool

        # Parsing assegnazioni docenti civics
        logging.info("Parsing delle assegnazioni dei docenti di educazione civica...")
        self.docenti_civics_classi = {}
        for _, row in self.docenti_civics_df.iterrows():
            nome_docente = row['DOCENTE']
            lista_classi = row['CLASSI'].split(';')
            self.docenti_civics_classi[nome_docente] = lista_classi

        # Identificazione docenti di civics anche in altre materie
        logging.info("Identificazione dei docenti di civics che insegnano altre materie nelle classi...")
        self.docenti_civics_organico = defaultdict(list)
        for _, row in self.classi_df.iterrows():
            nome_classe = row['CLASSE']
            for giorno in self.giorni_settimana:
                colonna_doc = f'DOC {giorno}'
                lista_docenti = row[colonna_doc].split(';')
                for docente in lista_docenti:
                    if docente in self.docenti_civics_classi:
                        self.docenti_civics_organico[nome_classe].append(docente)

        # Generazione degli slot disponibili (classe, data, giorno, ora, docente_sostituito)
        logging.info("Generazione degli slot disponibili...")
        self.slot_disponibili = []
        for nome_classe in self.classi_df['CLASSE']:
            for data in self.date_scolastiche:
                nome_giorno = self.mappa_giorni[data.weekday()]
                if nome_giorno in self.giorni_settimana:
                    orario_classe = self.orari_classi[nome_classe][nome_giorno]
                    for ora_idx, docente_in_classe in enumerate(orario_classe):
                        if docente_in_classe != '':
                            slot = {
                                'CLASSE': nome_classe,
                                'DATA': data,
                                'GIORNO': nome_giorno,
                                'ORA': ora_idx + 1,
                                'DOCENTE_SOSTITUITO': docente_in_classe,
                                'KEY': f"{nome_classe}_{data.strftime('%Y%m%d')}_{ora_idx + 1}"
                            }
                            self.slot_disponibili.append(slot)

        # Debug info
        print(f"Numero totale di slot disponibili: {len(self.slot_disponibili)}")
        print(f"Classi trovate: {self.classi_df['CLASSE'].tolist()}")
        print(f"Docenti civics: {list(self.docenti_civics_classi.keys())}")

    def genera_calendario(self):
        # Funzione principale che esegue l'algoritmo genetico, genera popolazione,
        # esegue crossover, mutazione, selezione e infine salva i risultati

        logging.info("Inizializzazione della popolazione...")
        self.initialize_population()

        if len(self.population) == 0:
            logging.error("Impossibile generare una popolazione iniziale valida.")
            return

        migliore_fitness = float('inf')
        migliore_individuo = None
        generazioni_senza_miglioramento = 0

        logging.info("Esecuzione dell'algoritmo genetico...")
        for generazione in range(self.num_generazioni):
            logging.info(f"Generazione {generazione + 1}/{self.num_generazioni}")

            # Aggiorna probabilità di mutazione ed elitismo in base alla mancata miglioria
            self.probabilita_mutazione = self.calcola_probabilita_mutazione(generazioni_senza_miglioramento)
            self.hyperparams['probabilita_mutazione'] = self.probabilita_mutazione

            self.evaluate_population()
            self.population.sort(key=lambda x: x['fitness'])

            num_elite = max(1, int(self.calcola_elitismo_rate(generazioni_senza_miglioramento) * self.popolazione_size))
            elite = self.population[:num_elite]

            # Controllo miglioramento
            if self.population[0]['fitness'] < migliore_fitness:
                migliore_fitness = self.population[0]['fitness']
                migliore_individuo = self.population[0]['individuo']
                generazioni_senza_miglioramento = 0
            else:
                generazioni_senza_miglioramento += 1

            # Early stopping se nessun miglioramento
            if generazioni_senza_miglioramento >= self.early_stopping_n:
                logging.info("Early stopping attivato.")
                break

            # Ricombinazione e mutazione per generare la nuova popolazione
            self.select_and_generate_new_population(elite)

            # -------------------------------------------
            # Salvataggio dei risultati della generazione corrente
            # -------------------------------------------
            generation_dir = os.path.join(self.cartella_output, f"generation_{generazione+1}")
            os.makedirs(generation_dir, exist_ok=True)

            best_individual = self.population[0]['individuo']
            best_calendario = self.create_calendario(best_individual)

            # Salva calendar.csv
            calendario_df = pd.DataFrame(best_calendario)
            calendario_df.to_csv(os.path.join(generation_dir, 'calendar.csv'), index=False)

            # Calcola e salva teachersLost.csv per questa generazione
            statistiche_classi = self.calcola_statistiche(best_calendario)
            statistiche_df = pd.DataFrame(statistiche_classi)
            statistiche_df.to_csv(os.path.join(generation_dir, 'teachersLost.csv'), index=False)

            # Genera i file Excel anche per la generazione intermedia
            genera_file_excel(best_calendario, self.classi_df, self.docenti_civics_df, generation_dir)

        logging.info("Migliore individuo trovato con fitness: {}".format(migliore_fitness))

        # -------------------------
        # Salvataggio finale
        # -------------------------
        calendario = self.create_calendario(migliore_individuo)

        if not os.path.exists(self.cartella_output):
            os.makedirs(self.cartella_output)

        logging.info("Salvataggio del calendario finale in calendar.csv...")
        calendario_df = pd.DataFrame(calendario)
        calendario_df.to_csv(os.path.join(self.cartella_output, 'calendar.csv'), index=False)

        logging.info("Salvataggio delle statistiche finali in teachersLost.csv...")
        statistiche_classi = self.calcola_statistiche(calendario)
        statistiche_df = pd.DataFrame(statistiche_classi)
        statistiche_df.to_csv(os.path.join(self.cartella_output, 'teachersLost.csv'), index=False)

        logging.info("Generazione dei file Excel finali...")
        genera_file_excel(calendario, self.classi_df, self.docenti_civics_df, self.cartella_output)
        logging.info("File Excel finali generati con successo!")

    def calcola_statistiche(self, calendario):
        # Calcola le statistiche per classe e docente (ore perse, totali e percentuale)
        statistiche_classi = []
        for classe in self.classi_df['CLASSE']:
            slots_classe = [s for s in self.slot_disponibili if s['CLASSE'] == classe]
            ore_perse_docente = defaultdict(int)
            ore_totali_docente = defaultdict(int)

            # Calcolo delle ore totali per docente
            for slot in slots_classe:
                docente = slot['DOCENTE_SOSTITUITO']
                ore_totali_docente[docente] += 1

            # Calcolo delle ore perse assegnate ad un docente di civics
            for entry in calendario:
                if entry['CLASSE'] == classe:
                    docente = entry['DOCENTE_SOSTITUITO']
                    ore_perse_docente[docente] += 1

            # Calcolo percentuali e costruzione del dict da aggiungere al DataFrame
            for docente in ore_totali_docente:
                ore_perse = ore_perse_docente.get(docente, 0)
                ore_totali = ore_totali_docente[docente]
                percentuale = (ore_perse / ore_totali * 100) if ore_totali > 0 else 0
                statistiche_classi.append({
                    'CLASSE': classe,
                    'DOCENTE': docente,
                    'ORE_PERSE': ore_perse,
                    'ORE_TOTALI': ore_totali,
                    'PERCENTUALE_ORE_PERSE': f'{percentuale:.2f}'
                })
        return statistiche_classi

    def initialize_population(self):
        # Generazione della popolazione iniziale con approcci diversi (greedy, batch, random)
        self.population = []
        tentativi = 0
        max_tentativi = self.popolazione_size * 100

        num_greedy = int(0.3 * self.popolazione_size)
        num_batch = int(0.3 * self.popolazione_size)
        num_random = self.popolazione_size - num_greedy - num_batch

        with multiprocessing.Pool(processes=self.num_cores) as pool:
            # Generazione con approccio greedy
            logging.info("Generazione popolazione iniziale con approccio greedy...")
            while len(self.population) < num_greedy and tentativi < max_tentativi:
                batch_size = min(num_greedy - len(self.population), self.num_cores)
                results = pool.map(genera_individuo_greedy_helper, [(self, None)] * batch_size)
                for individuo in results:
                    if individuo is not None:
                        self.population.append({'individuo': individuo})
                tentativi += batch_size

            # Generazione con approccio per fasce (batch)
            logging.info("Generazione popolazione iniziale con approccio per fasce...")
            while len(self.population) < num_greedy + num_batch and tentativi < max_tentativi:
                batch_size = min(num_batch - (len(self.population) - num_greedy), self.num_cores)
                results = pool.map(genera_individuo_batch_helper, [(self, None)] * batch_size)
                for individuo in results:
                    if individuo is not None:
                        self.population.append({'individuo': individuo})
                tentativi += batch_size

            # Generazione con approccio random
            logging.info("Generazione popolazione iniziale con approccio casuale...")
            while len(self.population) < self.popolazione_size and tentativi < max_tentativi:
                batch_size = min(self.popolazione_size - len(self.population), self.num_cores)
                results = pool.map(genera_individuo_random_helper, [(self, None)] * batch_size)
                for individuo in results:
                    if individuo is not None:
                        self.population.append({'individuo': individuo})
                tentativi += batch_size

    def evaluate_population(self):
        # Calcolo della fitness per ogni individuo della popolazione in parallelo
        with multiprocessing.Pool(processes=self.num_cores) as pool:
            fitness_results = pool.map(calcola_fitness_helper, [(self, ind['individuo']) for ind in self.population])
        for i, fit in enumerate(fitness_results):
            self.population[i]['fitness'] = fit

    def select_and_generate_new_population(self, elite):
        # Selezione e generazione nuova popolazione
        selected = self.selezione([ind['individuo'] for ind in self.population], [ind['fitness'] for ind in self.population])
        new_population = elite.copy()
        while len(new_population) < self.popolazione_size:
            genitore1 = random.choice(selected)
            genitore2 = random.choice(selected)
            if random.random() < self.probabilita_crossover:
                figlio = self.crossover(genitore1, genitore2)
            else:
                figlio = genitore1.copy()

            figlio = self.mutazione(figlio)

            if self.verifica_vincoli(figlio):
                new_population.append({'individuo': figlio})

        self.population = new_population

    def create_calendario(self, individuo):
        # Crea la lista di dizionari rappresentante il calendario dall'individuo
        calendario = []
        for slot_key, docente_civics in individuo.items():
            slot_info = next(s for s in self.slot_disponibili if s['KEY'] == slot_key)
            calendario.append({
                'CLASSE': slot_info['CLASSE'],
                'DATA': slot_info['DATA'].strftime('%d/%m/%Y'),
                'GIORNO': slot_info['GIORNO'],
                'ORA': slot_info['ORA'],
                'DOCENTE_CIVICS': docente_civics,
                'DOCENTE_SOSTITUITO': slot_info['DOCENTE_SOSTITUITO']
            })
        return calendario

    def genera_individuo_random(self, _):
        return self.genera_individuo_base(strategy='random')

    def genera_individuo_greedy(self, _):
        return self.genera_individuo_base(strategy='greedy')

    def genera_individuo_batch(self, _):
        return self.genera_individuo_base(strategy='batch')

    def genera_individuo_base(self, strategy='random'):
        # Genera un individuo con la strategia indicata (greedy, batch, random)
        individuo = {}
        ore_per_classe = defaultdict(int)
        ore_per_docente_data = defaultdict(lambda: defaultdict(int))
        ore_settimanali_classe = defaultdict(lambda: defaultdict(int))

        if strategy == 'greedy':
            slot_copia = sorted(self.slot_disponibili, key=lambda x: x['DATA'])
        elif strategy == 'batch':
            slot_copia = sorted(self.slot_disponibili, key=lambda x: (x['CLASSE'], x['DATA']))
        else:
            slot_copia = self.slot_disponibili.copy()
            random.shuffle(slot_copia)

        # Assegna docenti civics in base alla strategia
        for slot in slot_copia:
            nome_classe = slot['CLASSE']
            data = slot['DATA']
            settimana = data.isocalendar()[1]

            # Controlla limite di ore totali e settimanali
            if ore_per_classe[nome_classe] >= self.ore_tot_civics:
                continue
            if ore_settimanali_classe[nome_classe][settimana] >= 1:
                continue

            nome_giorno = slot['GIORNO']
            ora = slot['ORA']
            slot_key = slot['KEY']
            docente_sostituito = slot['DOCENTE_SOSTITUITO']

            # Trova docenti civics possibili
            docenti_possibili = []
            for docente_civics in self.docenti_civics_classi:
                if nome_classe in self.docenti_civics_classi[docente_civics]:
                    disponibile = False
                    if (docente_civics in self.docenti_civics_organico[nome_classe]) and (docente_civics == docente_sostituito):
                        # Se il docente civics insegna anche la materia e coincide con il docente sostituito
                        disponibile = True
                    else:
                        # Controlla disponibilità sul giorno e ora
                        if len(self.disponibilita_civics[docente_civics][nome_giorno]) >= ora and \
                            self.disponibilita_civics[docente_civics][nome_giorno][ora - 1]:
                            disponibile = True

                    # Controllo se il docente non insegna due ore nello stesso giorno alla stessa ora
                    if disponibile and ore_per_docente_data[docente_civics].get(data, 0) != ora:
                        docenti_possibili.append(docente_civics)

            if docenti_possibili:
                if strategy == 'greedy':
                    docente_assegnato = docenti_possibili[0]
                else:
                    docente_assegnato = random.choice(docenti_possibili)
                individuo[slot_key] = docente_assegnato
                ore_per_classe[nome_classe] += 1
                ore_settimanali_classe[nome_classe][settimana] += 1
                ore_per_docente_data[docente_assegnato][data] = ora

        print(f"Individuo generato per strategia '{strategy}': {len(individuo)} assegnazioni")

        if self.verifica_vincoli(individuo):
            return individuo
        else:
            return None

    def verifica_vincoli(self, individuo):
        # Verifica se l'individuo rispetta i vincoli (ore tot per classe e max 1 ora a settimana per classe)
        ore_per_classe = defaultdict(int)
        ore_settimanali_classe = defaultdict(lambda: defaultdict(int))

        for slot_key in individuo:
            slot_info = next(s for s in self.slot_disponibili if s['KEY'] == slot_key)
            nome_classe = slot_info['CLASSE']
            data = slot_info['DATA']
            settimana = data.isocalendar()[1]
            ore_per_classe[nome_classe] += 1
            ore_settimanali_classe[nome_classe][settimana] += 1

            if ore_settimanali_classe[nome_classe][settimana] > 1:
                return False

        # Tutte le classi devono avere esattamente ore_tot_civics ore
        if not all(ore_per_classe[nome_classe] == self.ore_tot_civics
                   for nome_classe in self.classi_df['CLASSE']):
            return False

        return True

    def calcola_fitness(self, individuo):
        # Calcola la fitness di un individuo, utilizzando diverse metriche
        # Minore è la fitness, migliore è l'individuo
        ore_settimanali_classe = defaultdict(lambda: defaultdict(int))
        for slot_key in individuo:
            slot_info = next(s for s in self.slot_disponibili if s['KEY'] == slot_key)
            nome_classe = slot_info['CLASSE']
            data = slot_info['DATA']
            settimana = data.isocalendar()[1]
            ore_settimanali_classe[nome_classe][settimana] += 1

        total_deviation = 0
        for classe in self.classi_df['CLASSE']:
            ore_per_settimana = ore_settimanali_classe.get(classe, {})
            deviations = [max(0, ore - 1) for ore in ore_per_settimana.values()]
            total_deviation += sum(deviations)

        variance_total = 0
        max_percentage_penalty = 0
        penalties_total = 0

        medium_intensity_penalty = 5
        high_intensity_penalty = 10
        low_intensity_penalty = 1

        # Penalità per docenti civics
        medium_intensity_penalty_civics_teacher = 10
        high_intensity_penalty_civics_teacher = 20
        low_intensity_penalty_civics_teacher = 0.5

        for classe in self.classi_df['CLASSE']:
            slots_classe = [s for s in self.slot_disponibili if s['CLASSE'] == classe]
            ore_perse_docente = defaultdict(int)
            ore_totali_docente = defaultdict(int)

            for slot in slots_classe:
                docente = slot['DOCENTE_SOSTITUITO']
                ore_totali_docente[docente] += 1

            for slot_key in individuo:
                slot_info = next(s for s in self.slot_disponibili if s['KEY'] == slot_key)
                if slot_info['CLASSE'] == classe:
                    docente = slot_info['DOCENTE_SOSTITUITO']
                    ore_perse_docente[docente] += 1

            total_teaching_hours = sum(ore_totali_docente.values())
            P = (self.ore_tot_civics / total_teaching_hours) * 100 if total_teaching_hours > 0 else 0

            percentuali = []
            for docente in ore_totali_docente:
                ore_totali = ore_totali_docente[docente]
                ore_perse = ore_perse_docente.get(docente, 0)
                percentuale_perse = (ore_perse / ore_totali) * 100 if ore_totali > 0 else 0
                percentuali.append(percentuale_perse)

                # Penalità in base alla percentuale di ore perse
                if percentuale_perse > 2 * P:
                    if docente in self.docenti_civics_organico[classe]:
                        penalties_total += high_intensity_penalty_civics_teacher
                    else:
                        penalties_total += high_intensity_penalty
                elif percentuale_perse > P:
                    if docente in self.docenti_civics_organico[classe]:
                        penalties_total += medium_intensity_penalty_civics_teacher
                    else:
                        penalties_total += medium_intensity_penalty
                elif percentuale_perse < 0.3 * P:
                    if docente in self.docenti_civics_organico[classe]:
                        penalties_total += low_intensity_penalty_civics_teacher
                    else:
                        penalties_total += low_intensity_penalty

                # Penalità per percentuali molto alte
                if docente in self.docenti_civics_organico[classe]:
                    if percentuale_perse > 5:  
                        max_percentage_penalty += (percentuale_perse - 5) * 10

            if percentuali:
                variance = np.var(percentuali)
                variance_total += variance

        total_fitness = total_deviation * 10 + variance_total * 5 + max_percentage_penalty + penalties_total
        return total_fitness

    def selezione(self, popolazione, fitness):
        # Selezione con ranking
        popolazione_fitness = list(zip(popolazione, fitness))
        popolazione_fitness.sort(key=lambda x: x[1])
        ranks = range(len(popolazione), 0, -1)
        total_rank = sum(ranks)
        selection_probs = [rank / total_rank for rank in ranks]
        popolazione_sorted = [ind for ind, fit in popolazione_fitness]
        selected = random.choices(popolazione_sorted, weights=selection_probs, k=len(popolazione))
        return selected

    def crossover(self, genitore1, genitore2):
        # Crossover: unisce parti di genitore1 e genitore2
        figlio = {}
        blocks = self.identify_blocks(genitore1, genitore2)
        for block in blocks:
            if random.random() < 0.5:
                figlio.update(block['genitore1'])
            else:
                figlio.update(block['genitore2'])
        return figlio

    def identify_blocks(self, genitore1, genitore2):
        # Identifica blocchi di chiavi da scambiare
        keys = list(genitore1.keys())
        random.shuffle(keys)
        blocks = []
        block_size = max(1, len(keys) // 10)
        for i in range(0, len(keys), block_size):
            block_keys = keys[i:i+block_size]
            block_gen1 = {k: genitore1[k] for k in block_keys}
            block_gen2 = {k: genitore2.get(k, genitore1[k]) for k in block_keys}
            blocks.append({'genitore1': block_gen1, 'genitore2': block_gen2})
        return blocks

    def mutazione(self, individuo):
        # Mutazione casuale: in alcuni slot cambia il docente assegnato
        keys = list(individuo.keys())
        for key in keys:
            if random.random() < self.probabilita_mutazione:
                slot_info = next(s for s in self.slot_disponibili if s['KEY'] == key)
                nome_classe = slot_info['CLASSE']
                data = slot_info['DATA']
                settimana = data.isocalendar()[1]
                nome_giorno = slot_info['GIORNO']
                ora = slot_info['ORA']
                docente_sostituito = slot_info['DOCENTE_SOSTITUITO']

                docenti_possibili = []
                for docente_civics in self.docenti_civics_classi:
                    if nome_classe in self.docenti_civics_classi[docente_civics]:
                        disponibile = False
                        if docente_civics in self.docenti_civics_organico[nome_classe]:
                            if self.allow_teacher_replace_self and docente_civics == docente_sostituito:
                                disponibile = True
                        else:
                            if len(self.disponibilita_civics[docente_civics][nome_giorno]) >= ora and \
                               self.disponibilita_civics[docente_civics][nome_giorno][ora - 1]:
                                disponibile = True

                        if disponibile:
                            docenti_possibili.append(docente_civics)

                if docenti_possibili:
                    individuo[key] = random.choice(docenti_possibili)
        return individuo


def genera_individuo_greedy_helper(args):
    self, _ = args
    return self.genera_individuo_greedy(_)

def genera_individuo_batch_helper(args):
    self, _ = args
    return self.genera_individuo_batch(_)

def genera_individuo_random_helper(args):
    self, _ = args
    return self.genera_individuo_random(_)

def calcola_fitness_helper(args):
    self, individuo = args
    return self.calcola_fitness(individuo)


if __name__ == "__main__":
    generator = CalendarioGenerator(
        num_varianti=1,
        data_inizio_str='15/10/2024',
        data_fine_str='10/06/2025',
        ore_tot_civics=30,
        cartella_output="CALENDARIO_GENERATO",
        num_generazioni=400,
        early_stopping_n=2,
        popolazione_size=500,
        probabilita_mutazione=0.5,
        probabilita_crossover=0.8,
        elitismo_rate=0.005,
        num_cores=10,
        allow_teacher_replace_self=False
    )
    generator.genera_calendario()

